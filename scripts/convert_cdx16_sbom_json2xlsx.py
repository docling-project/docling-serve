"""
convert_cdx16_sbom_json2xlsx.py
--------------------------------
Converts a CycloneDX 1.6 JSON SBOM file into an Excel workbook with four
sheets:

  1. Metadata       — SBOM timestamp, root component, and toolchain info
  2. SBOM Explanation — Human-readable description of the SBOM generation
                       process and how to interpret the workbook
  3. Components     — Sorted table of all software packages (name, version,
                       author, license, description, PURL, path, BOM ref,
                       hash)
  4. Dependencies   — Component dependency relationships

Usage:
    python scripts/convert_cdx16_sbom_json2xlsx.py -i <sbom.cdx.json> -o <output.xlsx>

Dependencies:
    pip install pandas openpyxl
"""

import argparse
import json
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font


def _format_tool_entry(tool: dict) -> str | None:
    """Format a single tool dict into a human-readable string, or None if not a dict."""
    if not isinstance(tool, dict):
        return None
    name = tool.get("name", "Unknown Tool")
    version = tool.get("version", "")
    group = tool.get("group", "")
    author = tool.get("author", "")
    license_ids = [
        lic.get("license", {}).get("id")
        for lic in tool.get("licenses", [])
        if lic.get("license", {}).get("id")
    ]
    parts = [name]
    if group:
        parts.append(f"(Group: {group})")
    if version:
        parts.append(f"v{version}")
    if author:
        parts.append(f"by {author}")
    if license_ids:
        parts.append(f"[License: {', '.join(license_ids)}]")
    return " ".join(parts)


def _flatten_component(component: dict, lookup: dict, parent_ref: str = "") -> None:
    """Recursively flatten a component and its nested components into lookup."""
    bom_ref = component.get("bom-ref") or (
        f"{parent_ref}|{component['name']}@{component.get('version', '')}"
    )
    lookup[bom_ref] = component
    for nested in component.get("components", []):
        _flatten_component(nested, lookup, bom_ref)


def _build_component_row(bom_ref: str, component: dict) -> dict:
    """Build a row dict for the Components sheet from a single component."""
    name = component.get("name", "")
    version = component.get("version", "")
    group = component.get("group", "")
    desc = component.get("description", "")
    full_name = f"{group + '/' if group else ''}{name}"

    licenses = component.get("licenses", [])
    if licenses:
        license_info = ", ".join(
            lic.get("license", {}).get("id", "")
            for lic in licenses
            if lic.get("license", {}).get("id")
        )
        license_info = license_info or "No Assertion"
    else:
        license_info = "No Assertion"

    author = component.get("author") or "Not Listed"
    purl = component.get("purl", "")

    hash_val = ""
    for extref in component.get("externalReferences", []):
        if extref.get("type") == "distribution":
            hashes = extref.get("hashes", [])
            if hashes:
                hash_val = f"{hashes[0]['alg']}:{hashes[0]['content']}"
                break

    path_prop = next(
        (
            p["value"]
            for p in component.get("properties", [])
            if p["name"].endswith("package:path")
        ),
        "",
    )

    return {
        "Package Name": full_name,
        "Version": version,
        "Author": author,
        "License": license_info,
        "Description": desc,
        "PURL": purl,
        "Path": path_prop,
        "BOM Reference": bom_ref,
        "Hash Value": hash_val,
    }


def sbom_to_excel(input_json: Path, output_excel: Path) -> None:
    with open(input_json, encoding="utf-8") as f:
        sbom = json.load(f)

    # ── Metadata sheet ────────────────────────────────────────────────────────
    metadata = sbom.get("metadata", {})
    metadata_rows = []

    metadata_rows.append(["Timestamp", metadata.get("timestamp", "Not Available")])
    metadata_rows.append(
        ["Component Name", metadata.get("component", {}).get("name", "Not Available")]
    )
    metadata_rows.append(
        [
            "Component Version",
            metadata.get("component", {}).get("version", "Not Available"),
        ]
    )
    metadata_rows.append(
        ["Component Type", metadata.get("component", {}).get("type", "Not Available")]
    )
    metadata_rows.append(
        [
            "Component BOM Ref",
            metadata.get("component", {}).get("bom-ref", "Not Available"),
        ]
    )

    tools_section = metadata.get("tools", {})
    tool_components = tools_section.get("components", [])
    tool_entries = []

    for tool in tool_components:
        entry = _format_tool_entry(tool)
        if entry is not None:
            tool_entries.append(entry)

    if tool_entries:
        metadata_rows.append(["Tools Used", tool_entries[0]])
        for tool_entry in tool_entries[1:]:
            metadata_rows.append(["", tool_entry])
    else:
        metadata_rows.append(["Tools Used", "None"])

    metadata_df = pd.DataFrame(metadata_rows)

    # ── SBOM Explanation sheet ────────────────────────────────────────────────
    explanation_rows = [
        ["SBOM and Vulnerability Report Explanation"],
        [""],
        ["SBOM Generation Process:"],
        [
            "This Software Bill of Materials (SBOM) was generated using the cyclonedx-bom "
            "tool for Python, which produces a CycloneDX-compliant JSON file (version 1.6)."
        ],
        ["The process involves:"],
        [
            "- Installing the project's Python dependencies using the uv package manager."
        ],
        [
            "- Scanning the resulting Python virtual environment using the cyclonedx-py "
            "CLI (from the cyclonedx-bom package) to capture all installed packages."
        ],
        [
            "- Generating a CycloneDX 1.6 JSON SBOM that includes metadata, components, "
            "and dependency relationships."
        ],
        [
            "- Merging the environment SBOM with the GitLab container-scanning SBOM using "
            "the CycloneDX CLI (cyclonedx-linux-x64 v0.29.1) to produce a unified view."
        ],
        [
            "- Converting the merged JSON SBOM into this Excel format for improved human "
            "readability and compliance tracking."
        ],
        [""],
        ["Two SBOMs are produced per pipeline run:"],
        [
            "- Build SBOM: All project dependencies including development and optional "
            "extras (uv sync --frozen --all-extras)."
        ],
        [
            "- Runtime SBOM: Production dependencies only — development group excluded "
            "(uv sync --frozen --no-dev)."
        ],
        [""],
        ["Tools Used:"],
        ["- uv — Python package manager used to install and lock dependencies."],
        [
            "- cyclonedx-bom — Python SBOM generator (cyclonedx-py environment subcommand)."
        ],
        ["- CycloneDX CLI v0.29.1 — Used to merge multiple SBOM files hierarchically."],
        [""],
        ["SBOM Structure:"],
        [
            "- Metadata Sheet: Contains information about the root component and the "
            "tools used to generate the SBOM."
        ],
        [
            "- Components Sheet: Lists all software packages, including name, version, "
            "author, license, and package URL (PURL)."
        ],
        [
            "- Dependencies Sheet: Shows the direct dependencies of each component, "
            "representing the structure of the software supply chain."
        ],
        [""],
        ["How to Interpret the SBOM:"],
        [
            "- Each row in the Components sheet represents a Python package (or OS-level "
            "package from the container scan) included in the application."
        ],
        [
            "- The License column reflects the open-source license(s) associated with "
            "each component."
        ],
        [
            "- The Dependencies sheet shows how components depend on each other, "
            "identifying both direct and transitive relationships."
        ],
        [
            "- The Metadata sheet provides context on the SBOM's origin and the toolchain "
            "used to produce it."
        ],
    ]
    df_explanation = pd.DataFrame(explanation_rows)

    # ── Components sheet ──────────────────────────────────────────────────────
    components = sbom.get("components", [])
    component_lookup: dict = {}
    for c in components:
        _flatten_component(c, component_lookup)

    rows = [
        _build_component_row(bom_ref, comp)
        for bom_ref, comp in component_lookup.items()
    ]

    components_df = pd.DataFrame(rows)
    components_df = components_df.sort_values(by=["Package Name", "Version"])

    # ── Dependencies sheet ────────────────────────────────────────────────────
    dep_rows = [
        {"Component": d.get("ref", ""), "Depends On": dep}
        for d in sbom.get("dependencies", [])
        for dep in d.get("dependsOn", [])
    ]
    dependencies_df = pd.DataFrame(dep_rows)

    # ── Write workbook ────────────────────────────────────────────────────────
    with pd.ExcelWriter(output_excel) as writer:
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False, header=False)
        df_explanation.to_excel(
            writer, sheet_name="SBOM Explanation", index=False, header=False
        )
        components_df.to_excel(writer, sheet_name="Components", index=False)
        dependencies_df.to_excel(writer, sheet_name="Dependencies", index=False)

    # Bold the label column in Metadata sheet
    wb = openpyxl.load_workbook(output_excel)
    ws = wb["Metadata"]
    bold = Font(bold=True)
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).font = bold
    wb.save(output_excel)

    print(f"Excel workbook created: {output_excel}")
    print(
        f"  Sheets: Metadata, SBOM Explanation, Components ({len(rows)} packages), "
        f"Dependencies ({len(dep_rows)} edges)"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert a CycloneDX 1.6 SBOM JSON file to an Excel workbook."
    )
    parser.add_argument(
        "--input", "-i", required=True, help="Path to input CycloneDX JSON SBOM file"
    )
    parser.add_argument(
        "--output",
        "-o",
        default="cyclonedx_report.xlsx",
        help="Path to output Excel file (default: cyclonedx_report.xlsx)",
    )
    args = parser.parse_args()

    input_file = Path(args.input)
    output_file = Path(args.output)

    if not input_file.exists():
        print(f"ERROR: Input file not found: {input_file}")
        raise SystemExit(1)

    sbom_to_excel(input_file, output_file)
